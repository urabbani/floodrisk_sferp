import os
import gc
import sys
import numpy as np
import pandas as pd
import geopandas as gpd
import rasterio
from rasterio.windows import Window
from rasterio.features import rasterize
from rasterio.warp import reproject, Resampling
from pathlib import Path
from datetime import datetime
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from concurrent.futures import ProcessPoolExecutor, as_completed

# GPU ACCELERATION
import cupy as cp

# Windows CUDA Path Fix
if "CUDA_PATH" not in os.environ:
    os.environ["CUDA_PATH"] = r"C:\Program Files\NVIDIA GPU Computing Toolkit\CUDA\v12.6"
cuda_bin = os.path.join(os.environ.get("CUDA_PATH", ""), "bin")
if os.path.exists(cuda_bin):
    os.add_dll_directory(cuda_bin)

# =============================================================================
# CONFIGURATION
# =============================================================================

DB_CONFIG = {
    "host": "10.0.0.205", "port": "5432", "database": "postgres",
    "user": "postgres", "password": "maltanadirSRV0", "schema_in": "Exposure_InputData"
}

DISTRICT_SHP_PATH = "Districts.shp" # Ensure this file exists in the script folder

ECONOMIC_VALUES = {
    "crop_per_ha": 2880.6,
    "build_low_56_per_sqm": 82.9,
    "build_low_44_per_sqm": 414.52,
    "build_high_per_sqm": 829.04
}

DATA_DIR_RASTER = "Exposure_InputData_32642"
IMPACT_THRESHOLD = 0.01 
BIN_EDGES = np.array([0.0, 0.5, 1.0, 1.5, 2.0, 2.5, 3.0, 3.5, 4.0, 4.5, 5.0, 5.5, 6.0, 6.5, 7.0, 7.5, 8.0, 8.5, 9.0, 9.5, 10.0, 1000.0])
DEPTH_BIN_LABELS_CM = [0, 50, 100, 150, 200, 250, 300, 350, 400, 450, 500, 550, 600, 650, 700, 750, 800, 850, 900, 950, 1000]

CHUNK_SIZE = 4096 

def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

# =============================================================================
# PROCESSING FUNCTIONS
# =============================================================================

def process_one_category(cat, flood_path, gdf=None, district_geom=None):
    try:
        t = cat['type']
        if t == 'polygon': return process_polygon_area(gdf, flood_path, district_geom)
        elif t == 'raster_area': return process_raster_area_gpu(cat, flood_path, district_geom)
        elif t == 'raster_height_area': return process_building_height_area_gpu(cat, flood_path, district_geom)
    except Exception as e: return f"Error: {str(e)}"

def process_polygon_area(gdf, flood_path, district_geom=None):
    results_m2 = np.zeros(21)
    with rasterio.open(flood_path) as src:
        pixel_area = abs(src.transform[0] * src.transform[4])
        target_gdf = gdf
        if district_geom: target_gdf = gdf[gdf.intersects(district_geom)]
        if target_gdf.empty: return {'ha': [0.0]*21}
        
        for row_off in range(0, src.height, CHUNK_SIZE):
            h = min(CHUNK_SIZE, src.height - row_off)
            for col_off in range(0, src.width, CHUNK_SIZE):
                w = min(CHUNK_SIZE, src.width - col_off); win = Window(col_off, row_off, w, h)
                trans = src.window_transform(win); bounds = src.window_bounds(win)
                chunk_gdf = target_gdf.cx[bounds[0]:bounds[2], bounds[1]:bounds[3]]
                if chunk_gdf.empty: continue
                mask = rasterize([(g, 1) for g in chunk_gdf.geom], out_shape=(h, w), transform=trans, fill=0)
                if district_geom:
                    d_mask = rasterize([(district_geom, 1)], out_shape=(h, w), transform=trans, fill=0)
                    mask = mask & d_mask
                f_chunk = src.read(1, window=win)
                px = f_chunk[(mask == 1) & (f_chunk > IMPACT_THRESHOLD)]
                if px.size > 0:
                    counts, _ = np.histogram(px, bins=BIN_EDGES)
                    results_m2 += counts * pixel_area
    return {'ha': (results_m2 / 10000.0).tolist()}

def process_raster_area_gpu(cat, flood_path, district_geom=None):
    results_m2 = np.zeros(21)
    cp_bins = cp.array(BIN_EDGES)
    with rasterio.open(flood_path) as f_src, rasterio.open(os.path.join(DATA_DIR_RASTER, cat['layer'])) as d_src:
        pixel_area = abs(f_src.transform[0] * f_src.transform[4])
        for row_off in range(0, f_src.height, CHUNK_SIZE):
            h = min(CHUNK_SIZE, f_src.height - row_off)
            for col_off in range(0, f_src.width, CHUNK_SIZE):
                w = min(CHUNK_SIZE, f_src.width - col_off); win = Window(col_off, row_off, w, h)
                f_chunk = cp.array(f_src.read(1, window=win)); d_chunk_cpu = np.zeros((h, w), dtype=np.float32)
                reproject(rasterio.band(d_src, 1), d_chunk_cpu, src_transform=d_src.transform, src_crs=d_src.crs, dst_transform=f_src.window_transform(win), dst_crs=f_src.crs, resampling=Resampling.nearest)
                d_chunk = cp.array(d_chunk_cpu)
                if district_geom:
                    d_mask = cp.array(rasterize([(district_geom, 1)], out_shape=(h, w), transform=f_src.window_transform(win), fill=0))
                    d_chunk *= d_mask
                for b_idx in range(21):
                    mask = (f_chunk >= cp_bins[b_idx]) & (f_chunk < cp_bins[b_idx+1]) & (d_chunk > 0)
                    results_m2[b_idx] += float(cp.count_nonzero(mask)) * pixel_area
    return {'m2': results_m2.tolist()}

def process_building_height_area_gpu(cat, flood_path, district_geom=None):
    res = {'<3m': np.zeros(21), '>=3m': np.zeros(21)}
    cp_bins = cp.array(BIN_EDGES)
    with rasterio.open(flood_path) as f_src, rasterio.open(os.path.join(DATA_DIR_RASTER, cat['layer'])) as h_src:
        pixel_area = abs(f_src.transform[0] * f_src.transform[4])
        for row_off in range(0, f_src.height, CHUNK_SIZE):
            h = min(CHUNK_SIZE, f_src.height - row_off)
            for col_off in range(0, f_src.width, CHUNK_SIZE):
                w = min(CHUNK_SIZE, f_src.width - col_off); win = Window(col_off, row_off, w, h)
                f_chunk = cp.array(f_src.read(1, window=win)); h_chunk_cpu = np.zeros((h, w), dtype=np.float32)
                reproject(rasterio.band(h_src, 1), h_chunk_cpu, src_transform=h_src.transform, src_crs=h_src.crs, dst_transform=f_src.window_transform(win), dst_crs=f_src.crs, resampling=Resampling.nearest)
                h_chunk = cp.array(h_chunk_cpu)
                if district_geom:
                    d_mask = cp.array(rasterize([(district_geom, 1)], out_shape=(h, w), transform=f_src.window_transform(win), fill=0))
                    h_chunk *= d_mask
                for b_idx in range(21):
                    mask_f = (f_chunk >= cp_bins[b_idx]) & (f_chunk < cp_bins[b_idx+1])
                    res['<3m'][b_idx] += float(cp.count_nonzero((h_chunk > 0) & (h_chunk < 3.0) & mask_f)) * pixel_area
                    res['>=3m'][b_idx] += float(cp.count_nonzero((h_chunk >= 3.0) & mask_f)) * pixel_area
    return res

# =============================================================================
# EXCEL GENERATION
# =============================================================================

def populate_sheet(ws, all_results, region_name, mode="Exp", multipliers=None):
    headers = ["Depth (cm)", "Crop (ha)", "Build <3m 56% (sqm)", "Build <3m 44% (sqm)", "Build >=3m (sqm)"]
    ws.append([f"Region: {region_name} | Mode: {mode}"])
    ws.append([])
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=text); cell.font = Font(bold=True)
    
    crop_ha = all_results.get("Cropped Area", {}).get('ha', [0.0]*21)
    build_low = all_results.get("Building Height", {}).get('<3m', [0.0]*21)
    build_high = all_results.get("Building Height", {}).get('>=3m', [0.0]*21)

    for i in range(21):
        c, bl56, bl44, bh = crop_ha[i], build_low[i]*0.56, build_low[i]*0.44, build_high[i]
        if multipliers is not None:
            m = multipliers.iloc[i]
            c *= m.iloc[1]; bl56 *= m.iloc[2]; bl44 *= m.iloc[3]; bh *= m.iloc[4]
        if mode == "Dmg":
            c *= ECONOMIC_VALUES['crop_per_ha']; bl56 *= ECONOMIC_VALUES['build_low_56_per_sqm']; bl44 *= ECONOMIC_VALUES['build_low_44_per_sqm']; bh *= ECONOMIC_VALUES['build_high_per_sqm']
        ws.append([DEPTH_BIN_LABELS_CM[i], round(c, 2), round(bl56, 2), round(bl44, 2), round(bh, 2)])
    
    # Totals
    l_row = ws.max_row + 1; ws.cell(row=l_row, column=1, value="TOTAL").font = Font(bold=True)
    for col in range(2, 6):
        let = chr(64+col); ws.cell(row=l_row, column=col, value=f"=SUM({let}4:{let}{l_row-1})").font = Font(bold=True)

# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":
    if len(sys.argv) < 2: sys.exit(1)
    PARENT_DIR = Path(sys.argv[1]).resolve()
    scenario_folders = sorted([d for d in PARENT_DIR.glob("T3_*") if d.is_dir()])
    VULN_FILE = Path("VulnerabilityMultiplier.xlsx")
    multipliers_df = pd.read_excel(VULN_FILE) if VULN_FILE.exists() else None
    
    # Districts & Vectors
    districts_gdf = gpd.read_file(DISTRICT_SHP_PATH).to_crs("EPSG:32642")
    engine = create_engine(f"postgresql://{DB_CONFIG['user']}:{DB_CONFIG['password']}@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}")
    CROP_GDF = gpd.read_postgis(f"SELECT geom FROM \"{DB_CONFIG['schema_in']}\".\"Cropped_Area\"", engine, geom_col='geom')
    engine.dispose()

    for scenario_path in scenario_folders:
        scenario_name = scenario_path.name; flood_raster = scenario_path / f"{scenario_name}_MaxDepth.tif"
        if not flood_raster.exists(): continue
        
        log(f">>> Scenario: {scenario_name}")
        wb = Workbook(); wb.remove(wb.active) # Clear default sheet
        
        all_regions = [("TOTAL", None)] + [(row['district'], row['geometry']) for _, row in districts_gdf.iterrows()]
        
        for name, geom in all_regions:
            log(f"--- Region: {name}")
            results = {}
            # We use 3 workers since we have 3 infrastructure categories
            with ProcessPoolExecutor(max_workers=3) as executor:
                cats = [
                    {"id": 1, "name": "Cropped Area", "type": "polygon", "gdf": CROP_GDF},
                    {"id": 2, "name": "Building Presence", "type": "raster_area", "layer": "buildings_presence_2023.tif"},
                    {"id": 3, "name": "Building Height", "type": "raster_height_area", "layer": "buildings_height_2023.tif"}
                ]
                futures = {executor.submit(process_one_category, cat, str(flood_raster), cat.get('gdf'), geom): cat['name'] for cat in cats}
                for f in as_completed(futures): results[futures[f]] = f.result()
            
            # Sheet names limited to 31 chars
            safe_name = str(name)[:25]
            populate_sheet(wb.create_sheet(f"{safe_name}_Exp"), results, name, "Exp")
            if multipliers_df is not None:
                populate_sheet(wb.create_sheet(f"{safe_name}_Vul"), results, name, "Vul", multipliers_df)
                populate_sheet(wb.create_sheet(f"{safe_name}_Dmg"), results, name, "Dmg", multipliers_df)

        report_path = scenario_path / "Exposure" / f"Risk_Analysis_{scenario_name}.xlsx"
        os.makedirs(report_path.parent, exist_ok=True); wb.save(report_path)
        log(f"Saved: {report_path}")